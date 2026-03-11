from __future__ import annotations

import os
import re
import unicodedata
from functools import lru_cache
from typing import Dict, List, Optional, Tuple


# ============================
# Normalization (backend-only)
# ============================


def _strip_accents(s: str) -> str:
    return "".join(
        ch
        for ch in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(ch)
    )


def _norm_text(s: str) -> str:
    """Deterministic normalization (trim, uppercase, remove diacritics).

    Also removes punctuation and collapses whitespace.
    """
    s = (s or "").strip()
    if not s:
        return ""
    s = _strip_accents(s)
    s = s.upper()
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _split_name(raw: str) -> Tuple[str, str]:
    """Return (surname, name) from common formats.

    Supports:
    - "SURNAME, NAME"
    - "NAME SURNAME"
    - Weird legacy: "CARSEN EDWARDS,CARSEN" -> surname=EDWARDS, name=CARSEN
    """
    s = _norm_text(raw)
    if not s:
        return ("", "")

    if "," in (raw or ""):
        # Use original to detect comma placement, but normalized parts.
        parts_raw = (raw or "").split(",", 1)
        left = _norm_text(parts_raw[0])
        right = _norm_text(parts_raw[1]) if len(parts_raw) > 1 else ""

        # left may contain both first+last; take last token as surname
        ltoks = left.split()
        surname = ltoks[-1] if ltoks else ""
        name = right
        return (surname, name)

    toks = s.split()
    if len(toks) == 1:
        return (toks[0], "")

    surname = toks[-1]
    name = " ".join(toks[:-1]).strip()
    return (surname, name)


def canonical_player_key(team_key: str | None, player_name: str | None) -> str:
    """Canonical key for deterministic mapping when ids are missing.

    Format: TEAM_KEY|SURNAME|NAME
    """
    from app.utils.canonical import canonical_team, canonical_team_from_code

    team_raw = (team_key or "").strip()
    t = ""
    if team_raw:
        t = canonical_team_from_code(team_raw) or canonical_team(team_raw) or team_raw
    t = _norm_text(t)

    surname, name = _split_name(player_name or "")
    if not t or not surname:
        return ""

    return f"{t}|{surname}|{name}".strip("|")


def _parse_positions(pos_raw: str | None) -> List[str]:
    s = _norm_text(pos_raw or "")
    if not s:
        return []

    # allow separators: '/', ',', ';', whitespace
    s = re.sub(r"[\\/;,]+", " ", s)
    toks = [t for t in s.split() if t]

    out: List[str] = []
    for t in toks:
        if t in {"PG", "SG", "SF", "PF", "C"} and t not in out:
            out.append(t)
    return out


# ============================
# Excel loading + lookups
# ============================


def _positions_xlsx_path() -> str:
    from app.services.json_loader import DataStore

    ds = DataStore.get()
    # Source-of-truth file (kept in repo data root)
    return os.path.join(ds.data_dir, "METADATA", "MASTER.XLSX")


@lru_cache(maxsize=1)
def _load_positions() -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """Load mapping from MASTER.XLSX.

    Returns:
      (by_player_id, by_canonical_key)
    """
    by_id: Dict[str, List[str]] = {}
    by_key: Dict[str, List[str]] = {}

    fp = _positions_xlsx_path()
    if not os.path.exists(fp):
        return by_id, by_key

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        # Dependency missing; fail safe (no positions)
        return by_id, by_key

    try:
        wb = load_workbook(fp, data_only=True)
    except Exception:
        return by_id, by_key

    # Expected columns: Player, Position, Player_ID, Team
    for ws in wb.worksheets:
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                player_raw = row[0] if len(row) > 0 else None
                pos_raw = row[1] if len(row) > 1 else None
                pid_raw = row[2] if len(row) > 2 else None
                team_raw = row[3] if len(row) > 3 else None

                positions = _parse_positions(str(pos_raw) if pos_raw is not None else "")
                if not positions:
                    continue

                pid = str(pid_raw).strip() if pid_raw is not None else ""
                if pid:
                    cur = by_id.get(pid, [])
                    for p in positions:
                        if p not in cur:
                            cur.append(p)
                    by_id[pid] = cur

                # Build canonical key fallback too (useful for future feeds w/out ids)
                key = canonical_player_key(str(team_raw or ""), str(player_raw or ""))
                if key:
                    curk = by_key.get(key, [])
                    for p in positions:
                        if p not in curk:
                            curk.append(p)
                    by_key[key] = curk
        except Exception:
            # Defensive: ignore a problematic sheet
            continue

    return by_id, by_key


def resolve_player_positions(
    player_id: str | None,
    team_key: str | None,
    player_name: str | None,
) -> Optional[List[str]]:
    by_id, by_key = _load_positions()

    pid = str(player_id).strip() if player_id is not None else ""
    if pid and pid in by_id:
        return by_id[pid]

    key = canonical_player_key(team_key, player_name)
    if key and key in by_key:
        return by_key[key]

    return None


def resolve_player_position(
    player_id: str | None,
    team_key: str | None,
    player_name: str | None,
) -> Optional[str]:
    positions = resolve_player_positions(player_id, team_key, player_name)
    if not positions:
        return None
    return positions[0]

import os
import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Any, Dict, List, Literal, Optional, Tuple

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


DeviationKind = Literal["standard", "percentage"]

_SHEET_RE = re.compile(r"^(?P<stat>.+)_(?P<kind>standard|percentage)$", re.IGNORECASE)

# Stats where LOWER opponent numbers indicate WEAKNESS instead of strength.
_REVERSED_COLOR_STATS = {
    "turnovers",
    "fouls",
    "foulsa",
    "blocks",
}


def _norm_stat(stat: str) -> str:
    return (stat or "").strip().lower().replace(" ", "")


def _sign(value: Optional[float]) -> str:
    if value is None or value == 0:
        return ""
    return "+" if value > 0 else "-"


def _color(stat: str, value: Optional[float]) -> str:
    if value is None or value == 0:
        return "neutral"

    s = _norm_stat(stat)
    reversed_logic = s in _REVERSED_COLOR_STATS

    if not reversed_logic:
        return "green" if value > 0 else "red"
    return "green" if value < 0 else "red"


@dataclass(frozen=True)
class OppDefenceConfig:
    xlsx_path: str


def _auto_detect_from_dir(dir_path: str) -> Optional[str]:
    if not dir_path or not os.path.isdir(dir_path):
        return None

    preferred = [
        os.path.join(dir_path, "basketstories_oppdef.xlsx"),
        os.path.join(dir_path, "opponents_defence.xlsx"),
        os.path.join(dir_path, "opponent_defence.xlsx"),
        os.path.join(dir_path, "opp_defence.xlsx"),
    ]
    for p in preferred:
        if os.path.isfile(p):
            return p

    try:
        files = [
            os.path.join(dir_path, f)
            for f in os.listdir(dir_path)
            if f.lower().endswith(".xlsx")
        ]
        files = [p for p in files if os.path.isfile(p)]
        if not files:
            return None
        files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return files[0]
    except Exception:
        return None


def _default_xlsx_path() -> str:
    env_path = os.environ.get("OPPONENT_DEFENCE_XLSX")
    if env_path and os.path.isfile(env_path):
        return env_path

    export_dir = r"C:\DEV\PROPDUNKER\EXPORT OPPONENTS DEFENCE"
    detected = _auto_detect_from_dir(export_dir)
    if detected:
        return detected

    return os.path.join(export_dir, "opponents_defence.xlsx")


def _read_sheet_table(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        team = ws.cell(r, 1).value
        if team is None or str(team).strip() == "":
            continue

        row: Dict[str, Any] = {"team": str(team).strip()}

        games = ws.cell(r, 2).value
        try:
            row["games"] = int(float(games)) if games is not None else None
        except Exception:
            row["games"] = None

        positions: Dict[str, Optional[float]] = {}
        for idx, col_name in enumerate(headers[2:], start=3):
            if not col_name:
                continue
            v = ws.cell(r, idx).value
            try:
                positions[col_name] = float(v) if v is not None else None
            except Exception:
                positions[col_name] = None

        row["positions"] = positions
        rows.append(row)

    return headers, rows


def _load_xlsx(path: str) -> Dict[str, Any]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    if not os.path.isfile(path):
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True)

    out: Dict[str, Any] = {
        "source": {"xlsx_path": path, "sheets": wb.sheetnames},
        "stats": {},
    }

    for sheet_name in wb.sheetnames:
        m = _SHEET_RE.match(sheet_name)
        if not m:
            continue

        stat = m.group("stat")
        kind: DeviationKind = m.group("kind").lower()  # type: ignore

        ws = wb[sheet_name]
        headers, rows = _read_sheet_table(ws)

        for row in rows:
            colored: Dict[str, Dict[str, Any]] = {}
            for pos, val in (row.get("positions") or {}).items():
                colored[pos] = {
                    "value": val,
                    "sign": _sign(val),
                    "color": _color(stat, val),
                }
            row["positions"] = colored

        out["stats"].setdefault(stat, {})[kind] = {
            "headers": headers,
            "rows": rows,
        }

    return out


@lru_cache(maxsize=8)
def load_opponent_defence_cached(path: str, mtime: float) -> Dict[str, Any]:
    return _load_xlsx(path)


def load_opponent_defence(
    xlsx_path: Optional[str] = None,
    kind: Literal["standard", "percentage", "both"] = "both",
    stat: Optional[str] = None,
) -> Dict[str, Any]:
    path = xlsx_path or _default_xlsx_path()
    mtime = os.path.getmtime(path)
    data = load_opponent_defence_cached(path, mtime)

    out = {"source": data.get("source"), "stats": {}}
    stats = data.get("stats") or {}

    for s_name, kinds in stats.items():
        if stat and _norm_stat(s_name) != _norm_stat(stat):
            continue

        if kind == "both":
            out["stats"][s_name] = kinds
        else:
            if kind in kinds:
                out["stats"][s_name] = {kind: kinds[kind]}

    return out

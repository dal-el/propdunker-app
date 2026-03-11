import os
import re
from dataclasses import dataclass
from functools import lru_cache
from typing import Any, Dict, List, Literal, Optional, Tuple

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore


DeviationKind = Literal["standard", "percentage"]

_SHEET_RE = re.compile(r"^(?P<stat>.+)_(?P<kind>standard|percentage)$", re.IGNORECASE)

_REVERSED_COLOR_STATS = {
    "turnovers",
    "to",
    "fouls",
    "foulsa",
    "blocks",
    "blk",
}

# Excel -> UI
_STAT_UI_MAP: Dict[str, str] = {
    "points": "PTS",
    "ftm": "FTM",
    "fta": "FTA",
    "fg2m": "2PM",
    "fg2a": "2PA",
    "fg3m": "3PM",
    "fg3a": "3PA",
    "totalrebs": "REB",
    "ofrebs": "OREB",
    "defrebs": "DREB",
    "assists": "AST",
    "turnovers": "TO",
    "steals": "STL",
    "blocks": "BLK",
    "fouls": "FOULS",
    "foulsa": "FOULS D",
}


def _norm_stat(stat: str) -> str:
    return (stat or "").strip().lower().replace(" ", "")


def _round2(v: Optional[float]) -> Optional[float]:
    if v is None:
        return None
    try:
        x = float(v)
    except Exception:
        return None
    if not (x == x):  # NaN
        return None
    return float(f"{x:.2f}")  # stable 2-dec representation


def _sign(value: Optional[float]) -> str:
    if value is None or value == 0:
        return ""
    return "+" if value > 0 else "-"


def _color(stat: str, value: Optional[float]) -> str:
    if value is None or value == 0:
        return "neutral"

    s = _norm_stat(stat)
    reversed_logic = s in _REVERSED_COLOR_STATS

    if not reversed_logic:
        return "green" if value > 0 else "red"
    return "green" if value < 0 else "red"


def _cell(stat_for_color: str, value: Optional[float]) -> Dict[str, Any]:
    v = _round2(value)
    return {"value": v, "sign": _sign(v), "color": _color(stat_for_color, v)}


@dataclass(frozen=True)
class OppDefenceConfig:
    xlsx_path: str


def _ui_key(raw_stat: str) -> str:
    k = _norm_stat(raw_stat)
    return _STAT_UI_MAP.get(k, raw_stat)


def _auto_detect_from_dir(dir_path: str) -> Optional[str]:
    if not dir_path or not os.path.isdir(dir_path):
        return None

    preferred = [
        os.path.join(dir_path, "basketstories_oppdef.xlsx"),
        os.path.join(dir_path, "opponents_defence.xlsx"),
        os.path.join(dir_path, "opponent_defence.xlsx"),
        os.path.join(dir_path, "opp_defence.xlsx"),
    ]
    for p in preferred:
        if os.path.isfile(p):
            return p

    try:
        files = [
            os.path.join(dir_path, f)
            for f in os.listdir(dir_path)
            if f.lower().endswith(".xlsx")
        ]
        files = [p for p in files if os.path.isfile(p)]
        if not files:
            return None
        files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return files[0]
    except Exception:
        return None


def _default_xlsx_path() -> str:
    env_path = os.environ.get("OPPONENT_DEFENCE_XLSX")
    if env_path and os.path.isfile(env_path):
        return env_path

    export_dir = r"C:\DEV\PROPDUNKER\EXPORT OPPONENTS DEFENCE"
    detected = _auto_detect_from_dir(export_dir)
    if detected:
        return detected

    return os.path.join(export_dir, "opponents_defence.xlsx")


def _read_sheet_table(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        team = ws.cell(r, 1).value
        if team is None or str(team).strip() == "":
            continue

        row: Dict[str, Any] = {"team": str(team).strip()}

        games = ws.cell(r, 2).value
        try:
            row["games"] = int(float(games)) if games is not None else None
        except Exception:
            row["games"] = None

        positions: Dict[str, Optional[float]] = {}
        for idx, col_name in enumerate(headers[2:], start=3):
            if not col_name:
                continue
            v = ws.cell(r, idx).value
            try:
                positions[col_name] = float(v) if v is not None else None
            except Exception:
                positions[col_name] = None

        row["positions"] = positions
        rows.append(row)

    return headers, rows


def _index_rows_by_team(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        t = str(r.get("team", "")).strip()
        if t:
            idx[t] = r
    return idx


def _decorate_table(stat_for_color: str, headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    out_rows: List[Dict[str, Any]] = []
    for row in rows:
        colored: Dict[str, Dict[str, Any]] = {}
        positions = row.get("positions") or {}
        for pos, val in positions.items():
            vv: Optional[float]
            try:
                vv = float(val) if val is not None else None
            except Exception:
                vv = None
            colored[pos] = _cell(stat_for_color, vv)
        out_rows.append({"team": row.get("team"), "games": row.get("games"), "positions": colored})
    return {"headers": headers, "rows": out_rows}


def _sum_tables(stat_name_for_color: str, tables: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not tables:
        return None

    base = tables[0]
    headers = list(base.get("headers") or [])
    base_rows = list(base.get("rows") or [])
    if not headers or not base_rows:
        return None

    idxs = [_index_rows_by_team(list(t.get("rows") or [])) for t in tables]
    pos_keys = [h for h in headers[2:] if h]

    out_rows: List[Dict[str, Any]] = []
    for team_row in base_rows:
        team = str(team_row.get("team", "")).strip()
        if not team:
            continue

        games = team_row.get("games")

        summed_positions: Dict[str, Dict[str, Any]] = {}
        for pos in pos_keys:
            vals: List[Optional[float]] = []
            for ti in idxs:
                r = ti.get(team)
                if not r:
                    vals.append(None)
                    continue
                p = (r.get("positions") or {}).get(pos)
                if isinstance(p, dict):
                    vals.append(p.get("value"))
                else:
                    vals.append(p if isinstance(p, (int, float)) else None)

            if all(v is None for v in vals):
                v_sum: Optional[float] = None
            else:
                v_sum = 0.0
                for v in vals:
                    if isinstance(v, (int, float)):
                        v_sum += float(v)

            summed_positions[pos] = _cell(stat_name_for_color, v_sum)

        out_rows.append({"team": team, "games": games, "positions": summed_positions})

    return {"headers": headers, "rows": out_rows}


def _load_xlsx(path: str) -> Dict[str, Any]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    if not os.path.isfile(path):
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(path, data_only=True)

    raw_stats: Dict[str, Dict[DeviationKind, Dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        m = _SHEET_RE.match(sheet_name)
        if not m:
            continue

        raw_stat = m.group("stat")
        kind: DeviationKind = m.group("kind").lower()  # type: ignore

        ws = wb[sheet_name]
        headers, rows = _read_sheet_table(ws)
        raw_stats.setdefault(raw_stat, {})[kind] = _decorate_table(raw_stat, headers, rows)

    ui_stats: Dict[str, Dict[DeviationKind, Dict[str, Any]]] = {}
    for raw_stat, kinds in raw_stats.items():
        ui_k = _ui_key(raw_stat)
        ui_stats.setdefault(ui_k, {})
        for k, table in kinds.items():
            ui_stats[ui_k][k] = table

    # Derived FGM/FGA
    for kind in ("standard", "percentage"):
        k: DeviationKind = kind  # type: ignore

        t_2pm = ui_stats.get("2PM", {}).get(k)
        t_3pm = ui_stats.get("3PM", {}).get(k)
        if t_2pm and t_3pm:
            summed = _sum_tables("FGM", [t_2pm, t_3pm])
            if summed:
                ui_stats.setdefault("FGM", {})[k] = summed

        t_2pa = ui_stats.get("2PA", {}).get(k)
        t_3pa = ui_stats.get("3PA", {}).get(k)
        if t_2pa and t_3pa:
            summed = _sum_tables("FGA", [t_2pa, t_3pa])
            if summed:
                ui_stats.setdefault("FGA", {})[k] = summed

    # Combos
    combos = {
        "PR": ["PTS", "REB"],
        "PA": ["PTS", "AST"],
        "RA": ["REB", "AST"],
        "PRA": ["PTS", "REB", "AST"],
        "PB": ["PTS", "BLK"],
        "PRB": ["PTS", "REB", "BLK"],
        "SB": ["STL", "BLK"],
    }

    for kind in ("standard", "percentage"):
        k: DeviationKind = kind  # type: ignore
        for combo_key, parts in combos.items():
            tables: List[Dict[str, Any]] = []
            ok = True
            for p in parts:
                t = ui_stats.get(p, {}).get(k)
                if not t:
                    ok = False
                    break
                tables.append(t)
            if not ok:
                continue
            summed = _sum_tables(combo_key, tables)
            if summed:
                ui_stats.setdefault(combo_key, {})[k] = summed

    return {
        "source": {"xlsx_path": path, "sheets": wb.sheetnames},
        "stats": ui_stats,
    }


@lru_cache(maxsize=8)
def load_opponent_defence_cached(path: str, mtime: float) -> Dict[str, Any]:
    return _load_xlsx(path)


def load_opponent_defence(
    xlsx_path: Optional[str] = None,
    kind: Literal["standard", "percentage", "both"] = "both",
    stat: Optional[str] = None,
) -> Dict[str, Any]:
    path = xlsx_path or _default_xlsx_path()
    mtime = os.path.getmtime(path)
    data = load_opponent_defence_cached(path, mtime)

    out = {"source": data.get("source"), "stats": {}}
    stats = data.get("stats") or {}

    for s_name, kinds in stats.items():
        if stat and _norm_stat(s_name) != _norm_stat(stat):
            continue

        if kind == "both":
            out["stats"][s_name] = kinds
        else:
            if kind in kinds:
                out["stats"][s_name] = {kind: kinds[kind]}

    return out
